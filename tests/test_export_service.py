from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.support import activity_factory as activity_service
from tests.support import runtime_state_fixture
from worktrace.services import database_maintenance_service, export_service
from worktrace.services.settings_service import (
    get_bool_setting,
    get_setting,
    set_setting,
)

pytestmark = [pytest.mark.db, pytest.mark.integration]


def test_excel_export_file_creation(temp_db, tmp_path):
    aid = activity_service.create_activity(
        "Word", "word.exe", "Doc", start_time="2026-06-18 09:00:00"
    )
    activity_service.close_activity(aid, "2026-06-18 09:30:00")
    path = export_service.export_excel(
        "2026-06-18",
        "2026-06-18",
        str(tmp_path / "out.xlsx"),
    )
    assert Path(path).exists()
    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
    assert "Sessions" in wb.sheetnames
    assert "Activity Logs" not in wb.sheetnames
    headers = [cell.value for cell in wb["Sessions"][1]]
    assert headers == [
        "日期",
        "开始时间",
        "结束时间",
        "时长",
        "时长秒数",
        "项目",
        "状态",
        "备注",
        "修正时长",
        "是否已修正",
    ]


def test_exports_prefer_activity_file_name_for_wps_activity(temp_db, tmp_path):
    aid = activity_service.create_activity(
        "WPS Writer",
        "wps.exe",
        "合同审查意见.docx - WPS",
        file_path_hint="D:\\ClientA\\合同审查意见.docx",
        start_time="2026-06-18 09:00:00",
    )
    activity_service.finalize_created_activity(aid)
    activity_service.close_activity(aid, "2026-06-18 09:30:00")

    xlsx_path = export_service.export_excel(
        "2026-06-18",
        "2026-06-18",
        str(tmp_path / "out.xlsx"),
    )
    ws = load_workbook(xlsx_path)["Sessions"]
    headers = [cell.value for cell in ws[1]]
    assert "资源名称" not in headers
    assert "路径" not in headers


def test_export_all_and_clear_requires_confirmation(temp_db, tmp_path):
    path = export_service.export_all_local_data(str(tmp_path / "all.xlsx"))
    assert Path(path).exists()
    wb = load_workbook(path)
    assert "folder_project_rule" in wb.sheetnames
    assert "folder_rule_index_state" not in wb.sheetnames
    assert "folder_rule_file_index" not in wb.sheetnames
    assert "activity_resource" in wb.sheetnames
    with pytest.raises(ValueError):
        export_service.clear_all_local_data(confirm=False)


def _seed_business_data() -> int:
    aid = activity_service.create_activity(
        "Word", "word.exe", "Doc", start_time="2026-06-18 09:00:00"
    )
    activity_service.close_activity(aid, "2026-06-18 09:30:00")
    return aid


def test_clear_all_confirm_false_does_not_reset_db(temp_db) -> None:
    aid = _seed_business_data()
    with pytest.raises(ValueError):
        export_service.clear_all_local_data(confirm=False)
    activities = activity_service.get_activities_by_range(
        "2026-06-18", "2026-06-18"
    )
    assert any(activity["id"] == aid for activity in activities)


def test_clear_all_success_restores_user_intent_and_clears_runtime(temp_db) -> None:
    _seed_business_data()
    set_setting("user_paused", "true")
    set_setting("collector_status", "paused")
    runtime_state_fixture.set_setting("current_activity_snapshot", '{"app":"Word"}')

    export_service.clear_all_local_data(confirm=True)

    assert get_bool_setting("user_paused", False) is True
    assert get_setting("collector_status", "") == "paused"
    assert runtime_state_fixture.get_setting("current_activity_snapshot", "") == ""
    assert activity_service.get_activities_by_range(
        "2026-06-18", "2026-06-18"
    ) == []


def test_clear_all_success_reseeds_system_projects(temp_db) -> None:
    _seed_business_data()
    export_service.clear_all_local_data(confirm=True)
    from worktrace.db import get_connection

    with get_connection() as conn:
        rows = conn.execute(
            "SELECT name FROM project WHERE created_by = 'system'"
        ).fetchall()
    assert {row[0] for row in rows}


def test_clear_all_rejects_when_another_maintenance_operation_owns_gate(
    temp_db,
) -> None:
    aid = _seed_business_data()
    with database_maintenance_service.consistent_snapshot("competing_operation"):
        with pytest.raises(ValueError, match="operation_in_progress"):
            export_service.clear_all_local_data(confirm=True)

    activities = activity_service.get_activities_by_range(
        "2026-06-18", "2026-06-18"
    )
    assert any(activity["id"] == aid for activity in activities)


def test_clear_all_failure_rolls_back_data_and_fails_closed(
    temp_db,
    monkeypatch,
) -> None:
    aid = _seed_business_data()
    set_setting("user_paused", "false")
    set_setting("collector_status", "running")
    runtime_state_fixture.set_setting(
        "current_activity_snapshot",
        '{"persisted_activity_id":77}',
    )

    def fail_seed(_conn) -> None:
        raise RuntimeError("post-clear failure")

    monkeypatch.setattr(database_maintenance_service, "seed_defaults", fail_seed)

    with pytest.raises(RuntimeError, match="post-clear failure"):
        export_service.clear_all_local_data(confirm=True)

    assert database_maintenance_service.is_maintenance_in_progress() is True
    assert (
        database_maintenance_service.MAINTENANCE_COORDINATOR.phase
        is database_maintenance_service.MaintenancePhase.FAILED_CLOSED
    )
    assert get_bool_setting("user_paused", False) is True
    assert get_setting("collector_status", "") == "paused"
    assert runtime_state_fixture.get_setting("current_activity_snapshot", "") == ""
    activities = activity_service.get_activities_by_range(
        "2026-06-18", "2026-06-18"
    )
    assert any(item["id"] == aid for item in activities)


def test_maintenance_context_clears_runtime_state_without_legacy_settings(
    temp_db,
) -> None:
    runtime_state_fixture.set_setting(
        "current_activity_snapshot",
        '{"status":"normal"}',
    )

    with database_maintenance_service.consistent_snapshot("runtime_state_contract"):
        assert database_maintenance_service.is_maintenance_in_progress() is True
        assert runtime_state_fixture.get_setting(
            "current_activity_snapshot", ""
        ) == ""


def test_clear_all_success_invalidates_context_recompute_cache(temp_db) -> None:
    _seed_business_data()
    export_service.clear_all_local_data(confirm=True)
    assert activity_service.get_activities_by_range(
        "2026-06-18", "2026-06-18"
    ) == []
