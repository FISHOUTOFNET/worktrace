"""Service-level project attribution display contract tests.

These tests verify that official display attribution and report-visible
attribution stay separated across Timeline / Statistics / Export /
Details / Live Display. The policy module
(``project_attribution_policy``) is the single source of truth.

Contract summary:
- Official sources (manual / keyword_rule / folder_rule) surface the
  real project name in official display fields.
- Concrete context sources are report-visible, but not official.
- Candidate / uncategorized sources resolve to ``UNCATEGORIZED_PROJECT``.
- ``_is_project_anchor`` requires ``is_official_project is True``.
- Statistics ``by_project`` / ``project_duration_seconds`` count
  report-visible projects.
- CSV ``项目`` cell never falls back to raw ``project_name``.
- ``advance_ownership`` never confirms a non-official candidate.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta

import pytest

from worktrace.constants import (
    STATUS_ERROR,
    STATUS_EXCLUDED,
    STATUS_IDLE,
    STATUS_NORMAL,
    STATUS_PAUSED,
    UNCATEGORIZED_PROJECT,
)
from worktrace.db import dict_rows, get_connection, now_str
from worktrace.formatters import format_activity_project_cell
from worktrace.services import (
    activity_service,
    export_service,
    folder_rule_service,
    project_service,
    rule_service,
    settings_service,
    statistics_service,
    timeline_service,
    view_model_service,
)
from worktrace.services.project_inference_service import (
    assign_project_for_activity,
    invalidate_keyword_rule_cache,
)
from worktrace.services.project_ownership_service import (
    ProjectLabel,
    ProjectOwnershipState,
    ProjectTransition,
    advance_ownership,
    begin_ownership_for_new_resource,
)

pytestmark = [pytest.mark.contract, pytest.mark.integration, pytest.mark.db]

TODAY = date.today().isoformat()


def _create_activity(
    app="Word",
    process="winword.exe",
    title="A.docx",
    start="09:00:00",
    end="09:10:00",
    project_id=None,
    status=STATUS_NORMAL,
    file_path_hint=None,
):
    """Create a closed activity for TODAY."""
    activity_service.close_all_open_rows(f"{TODAY} {start}")
    aid = activity_service.create_activity(
        app,
        process,
        title,
        start_time=f"{TODAY} {start}",
        project_id=project_id,
        status=status,
        file_path_hint=file_path_hint,
    )
    activity_service.finalize_created_activity(aid)
    if end:
        activity_service.close_activity(aid, f"{TODAY} {end}")
    return aid


def _set_assignment_source(activity_id, source, project_id=None, suggested_name=None, confidence=80):
    """Directly set the assignment source for an activity.

    This bypasses the inference pipeline so we can test each source
    type independently against the policy contract. Sets
    Sets ``is_manual=1`` on the assignment so context recompute does not
    overwrite this explicit fixture projection. It does not mutate raw
    ``activity_log`` project fields.
    """
    ts = now_str()
    with get_connection() as conn:
        if project_id is None:
            project_id = _get_uncategorized_id(conn)
        is_manual = 1
        conn.execute(
            """
            INSERT INTO activity_project_assignment(
                activity_id, project_id, confidence, source, is_manual,
                suggested_project_name, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(activity_id) DO UPDATE SET
                project_id = excluded.project_id,
                confidence = excluded.confidence,
                source = excluded.source,
                is_manual = excluded.is_manual,
                suggested_project_name = excluded.suggested_project_name,
                updated_at = excluded.updated_at
            """,
            (activity_id, project_id, confidence, source, is_manual, suggested_name, ts, ts),
        )


def _get_uncategorized_id(conn) -> int:
    return project_service.get_or_create_uncategorized_project()


def _uncategorized_id() -> int:
    return project_service.get_or_create_uncategorized_project()


def test_manual_assignment_displays_as_official(temp_db):
    """manual source → official project in Timeline / Stats / Export."""
    pid = project_service.create_project("ManualProject")
    aid = _create_activity(start="09:00:00", end="09:10:00")
    _set_assignment_source(aid, "manual", project_id=pid, confidence=100)

    # Timeline
    sessions = timeline_service.get_project_sessions_by_date(TODAY)
    assert len(sessions) == 1
    assert sessions[0]["project_name"] == "ManualProject"
    assert sessions[0]["is_official_project"] is True
    assert sessions[0]["is_uncategorized"] is False

    # Details
    details = timeline_service.get_session_activity_details([aid], TODAY)
    assert details[0]["project_name"] == "ManualProject"
    assert details[0]["is_official_project"] is True

    # Statistics
    summary = statistics_service.get_statistics_export_summary(TODAY, TODAY)
    by_project_names = [g["display_name"] for g in summary["by_project"]]
    assert "ManualProject" in by_project_names
    assert summary["project_duration_seconds"] == 600

    # Export CSV cell
    rows = export_service.build_statistics_csv_rows(TODAY, TODAY)
    assert len(rows) == 1
    assert rows[0]["project"] == "ManualProject"


def test_keyword_rule_assignment_displays_as_official(temp_db):
    """keyword_rule (user-created) → official project."""
    pid = project_service.create_project("KeywordProject")
    rule_service.create_rule("contract", pid)
    invalidate_keyword_rule_cache()

    aid = _create_activity(
        app="Word",
        title="ContractReview.docx - Word",
        start="09:00:00",
        end="09:10:00",
    )
    assign_project_for_activity(aid)

    sessions = timeline_service.get_project_sessions_by_date(TODAY)
    assert sessions[0]["project_name"] == "KeywordProject"
    assert sessions[0]["is_official_project"] is True

    summary = statistics_service.get_statistics_export_summary(TODAY, TODAY)
    by_project_names = [g["display_name"] for g in summary["by_project"]]
    assert "KeywordProject" in by_project_names


def test_folder_rule_assignment_displays_as_official(temp_db):
    """folder_rule (user-created) → official project."""
    pid = project_service.create_project("FolderProject")
    folder_rule_service.create_or_update_folder_rule("D:\\FolderProject", pid)

    aid = _create_activity(
        app="Visual Studio Code",
        process="Code.exe",
        title="main.py - Visual Studio Code",
        start="09:00:00",
        end="09:10:00",
        file_path_hint="D:\\FolderProject\\main.py",
    )
    assign_project_for_activity(aid)

    sessions = timeline_service.get_project_sessions_by_date(TODAY)
    assert sessions[0]["project_name"] == "FolderProject"
    assert sessions[0]["is_official_project"] is True


def test_system_created_keyword_rule_does_not_become_official(temp_db):
    """project_rule.created_by='system' → NOT official."""
    pid = project_service.create_project("SystemRuleProject")
    # Insert a system-created keyword rule directly
    ts = now_str()
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO project_rule(project_id, rule_type, pattern, enabled, created_by, created_at, updated_at)
            VALUES (?, 'keyword', 'sysonly', 1, 'system', ?, ?)
            """,
            (pid, ts, ts),
        )
    invalidate_keyword_rule_cache()

    aid = _create_activity(title="sysonly review.docx", start="09:00:00", end="09:10:00")
    assign_project_for_activity(aid)

    sessions = timeline_service.get_project_sessions_by_date(TODAY)
    # The system keyword rule must NOT produce an official project
    assert sessions[0]["project_name"] == UNCATEGORIZED_PROJECT
    assert sessions[0]["is_official_project"] is False

    summary = statistics_service.get_statistics_export_summary(TODAY, TODAY)
    by_project_names = [g["display_name"] for g in summary["by_project"]]
    assert "SystemRuleProject" not in by_project_names


def test_suggested_project_name_displays_as_uncategorized(temp_db):
    """suggested_project_name → formal project = uncategorized,
    candidate_project_name retains the suggestion."""
    aid = _create_activity(start="09:00:00", end="09:10:00")
    _set_assignment_source(
        aid,
        "suggested_project_name",
        project_id=_uncategorized_id(),
        suggested_name="SuggestedClient",
        confidence=50,
    )

    sessions = timeline_service.get_project_sessions_by_date(TODAY)
    assert len(sessions) == 1
    assert sessions[0]["project_name"] == UNCATEGORIZED_PROJECT
    assert sessions[0]["is_official_project"] is False
    assert sessions[0]["is_uncategorized"] is True

    # Report rows carry candidate_project_name
    rows = timeline_service.get_report_activity_rows(TODAY, TODAY)
    assert rows[0]["candidate_project_name"] == "SuggestedClient"
    assert rows[0]["is_official_project"] is False
    assert rows[0]["display_project_name"] == UNCATEGORIZED_PROJECT

    # Statistics: not in by_project
    summary = statistics_service.get_statistics_export_summary(TODAY, TODAY)
    by_project_names = [g["display_name"] for g in summary["by_project"]]
    assert "SuggestedClient" not in by_project_names
    assert summary["project_duration_seconds"] == 0

    # Export: CSV cell = uncategorized
    csv_rows = export_service.build_statistics_csv_rows(TODAY, TODAY)
    assert csv_rows[0]["project"] == UNCATEGORIZED_PROJECT


def test_suggested_candidate_short_interrupt_not_report_merged(temp_db, tmp_path):
    """suggested_project_name stays candidate metadata across short merge."""
    project_id = project_service.create_project("ProjectA")
    first_a = _create_activity(app="Word", title="A1.docx", start="09:00:00", end="09:10:00")
    candidate_b = _create_activity(app="Edge", title="Suggested", start="09:10:00", end="09:11:00")
    second_a = _create_activity(app="Word", title="A2.docx", start="09:11:00", end="09:20:00")
    _set_assignment_source(first_a, "keyword_rule", project_id=project_id)
    _set_assignment_source(
        candidate_b,
        "suggested_project_name",
        project_id=_uncategorized_id(),
        suggested_name="SuggestedClient",
        confidence=50,
    )
    _set_assignment_source(second_a, "keyword_rule", project_id=project_id)

    rows = timeline_service.get_report_activity_rows(TODAY, TODAY)
    b_row = next(row for row in rows if int(row["id"]) == candidate_b)
    assert b_row["candidate_project_name"] == "SuggestedClient"
    assert b_row["is_report_project"] is False
    assert b_row["report_project_name"] == UNCATEGORIZED_PROJECT
    assert b_row["report_attribution_kind"] == "candidate"
    assert b_row["report_context_merged"] is False

    sessions = timeline_service.get_project_sessions_by_date(TODAY)
    assert len(sessions) == 3
    b_session = next(row for row in sessions if row["start_time"].endswith("09:10:00"))
    assert b_session["project_name"] == UNCATEGORIZED_PROJECT
    assert b_session["duration_seconds"] == 60
    assert b_session["is_report_project"] is False
    assert b_session["report_attribution_kind"] == "candidate"
    project_a_seconds = sum(
        int(row["duration_seconds"])
        for row in sessions
        if row["project_name"] == "ProjectA"
    )
    assert project_a_seconds == 19 * 60

    summary = statistics_service.get_statistics_export_summary(TODAY, TODAY)
    assert summary["project_duration_seconds"] == 19 * 60
    assert [group["display_name"] for group in summary["by_project"]] == ["ProjectA"]

    csv_rows = export_service.build_statistics_csv_rows(TODAY, TODAY)
    b_csv = next(row for row in csv_rows if row["start_time"].endswith("09:10:00"))
    assert b_csv["project"] == UNCATEGORIZED_PROJECT

    from openpyxl import load_workbook

    xlsx_path = export_service.export_excel(TODAY, TODAY, str(tmp_path / "report.xlsx"))
    ws = load_workbook(xlsx_path)["Sessions"]
    headers = [cell.value for cell in ws[1]]
    project_col = headers.index("项目") + 1
    start_col = headers.index("开始时间") + 1
    excel_rows = list(ws.iter_rows(min_row=2, values_only=True))
    b_excel = next(row for row in excel_rows if str(row[start_col - 1]).endswith("09:10:00"))
    assert b_excel[project_col - 1] == UNCATEGORIZED_PROJECT


def test_suggested_project_name_does_not_confirm_after_threshold(temp_db):
    """advance_ownership must NOT confirm a suggested candidate."""
    candidate = ProjectLabel(
        name="SuggestedClient",
        id=None,
        source="suggested_project_name",
        is_suggested_project=True,
    )
    state = begin_ownership_for_new_resource(None, candidate, "2026-07-01 09:00:00")
    # Display should be uncategorized, not the candidate
    assert state.display_project.name == UNCATEGORIZED_PROJECT
    assert state.last_confirmed_project is None
    assert state.project_transition.pending is False

    # Advance 40 seconds (past the 30s threshold)
    state = advance_ownership(state, "2026-07-01 09:00:40")
    # Still NOT confirmed
    assert state.display_project.name == UNCATEGORIZED_PROJECT
    assert state.last_confirmed_project is None


@pytest.mark.parametrize(
    "source",
    ["anchor_context", "same_project_context", "clipboard_transition_context", "midnight_anchor"],
)
def test_context_sources_are_report_visible_but_not_official(temp_db, source):
    """Context sources stay unofficial while surfacing in report fields."""
    pid = project_service.create_project("InternalProject")
    aid = _create_activity(start="09:00:00", end="09:10:00")
    _set_assignment_source(aid, source, project_id=pid, confidence=60)

    sessions = timeline_service.get_project_sessions_by_date(TODAY)
    assert sessions[0]["project_name"] == "InternalProject"
    assert sessions[0]["is_official_project"] is False
    assert sessions[0]["is_report_project"] is True

    # Internal effective project id is preserved
    rows = timeline_service.get_report_activity_rows(TODAY, TODAY)
    assert rows[0]["effective_project_id"] == pid
    assert rows[0]["display_project_name"] == UNCATEGORIZED_PROJECT
    assert rows[0]["report_project_name"] == "InternalProject"
    assert rows[0]["is_official_project"] is False
    assert rows[0]["is_report_project"] is True

    summary = statistics_service.get_statistics_export_summary(TODAY, TODAY)
    by_project_names = [g["display_name"] for g in summary["by_project"]]
    assert "InternalProject" in by_project_names
    assert summary["project_duration_seconds"] == 600

    csv_rows = export_service.build_statistics_csv_rows(TODAY, TODAY)
    assert csv_rows[0]["project"] == "InternalProject"


def test_context_assignment_reports_across_timeline_details_stats_csv_excel(temp_db, tmp_path):
    project_id = project_service.create_project("ProjectA")
    first_a = _create_activity(app="Word", title="A1.docx", start="09:00:00", end="09:10:00")
    middle_b = _create_activity(app="Edge", title="Reference", start="09:10:00", end="09:11:00")
    second_a = _create_activity(app="Word", title="A2.docx", start="09:11:00", end="09:20:00")
    _set_assignment_source(first_a, "keyword_rule", project_id=project_id)
    _set_assignment_source(middle_b, "same_project_context", project_id=project_id)
    _set_assignment_source(second_a, "keyword_rule", project_id=project_id)

    rows = timeline_service.get_report_activity_rows(TODAY, TODAY)
    b_row = next(row for row in rows if int(row["id"]) == middle_b)
    assert b_row["display_project_name"] == UNCATEGORIZED_PROJECT
    assert b_row["report_project_name"] == "ProjectA"
    assert b_row["is_official_project"] is False
    assert b_row["is_report_project"] is True

    sessions = timeline_service.get_project_sessions_by_date(TODAY)
    assert len(sessions) == 1
    assert sessions[0]["project_name"] == "ProjectA"
    assert sessions[0]["is_report_project"] is True

    details = timeline_service.get_session_activity_details([middle_b], TODAY)
    assert details[0]["project_name"] == "ProjectA"
    assert details[0]["official_project_name"] == UNCATEGORIZED_PROJECT
    assert details[0]["is_official_project"] is False
    assert details[0]["is_report_project"] is True

    summary = statistics_service.get_statistics_export_summary(TODAY, TODAY)
    assert summary["project_duration_seconds"] == 20 * 60
    assert [group["display_name"] for group in summary["by_project"]] == ["ProjectA"]

    csv_rows = export_service.build_statistics_csv_rows(TODAY, TODAY)
    assert len(csv_rows) == 1
    assert csv_rows[0]["start_time"].endswith("09:00:00")
    assert csv_rows[0]["project"] == "ProjectA"

    from openpyxl import load_workbook

    xlsx_path = export_service.export_excel(TODAY, TODAY, str(tmp_path / "report.xlsx"))
    ws = load_workbook(xlsx_path)["Sessions"]
    headers = [cell.value for cell in ws[1]]
    project_col = headers.index("项目") + 1
    start_col = headers.index("开始时间") + 1
    excel_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(excel_rows) == 1
    assert str(excel_rows[0][start_col - 1]).endswith("09:00:00")
    assert excel_rows[0][project_col - 1] == "ProjectA"


@pytest.mark.parametrize(
    "source",
    ["anchor_context", "same_project_context", "clipboard_transition_context", "midnight_anchor"],
)
def test_derived_internal_sources_are_not_project_anchors(temp_db, source):
    """_is_project_anchor returns False for derived sources."""
    from worktrace.services.timeline_service import _is_project_anchor

    pid = project_service.create_project("InternalProject")
    aid = _create_activity(
        app="Word",
        title="Report.docx - Word",
        start="09:00:00",
        end="09:10:00",
        file_path_hint="D:\\Case\\Report.docx",
    )
    _set_assignment_source(aid, source, project_id=pid)

    rows = timeline_service.get_report_activity_rows(TODAY, TODAY)
    assert len(rows) == 1
    assert _is_project_anchor(rows[0]) is False


def test_format_activity_project_cell_no_raw_project_name_fallback():
    """format_activity_project_cell must NOT fall back to raw project_name."""
    # A row with a raw project_name but is_official_project=False
    row = {
        "status": STATUS_NORMAL,
        "project_name": "LeakedProject",
        "is_official_project": False,
    }
    assert format_activity_project_cell(row) == UNCATEGORIZED_PROJECT


def test_format_activity_project_cell_official_shows_name():
    row = {
        "status": STATUS_NORMAL,
        "report_project_name": "OfficialProject",
        "is_official_project": True,
        "is_report_project": True,
    }
    assert format_activity_project_cell(row) == "OfficialProject"


def test_format_activity_project_cell_non_normal_returns_dash():
    for status in [STATUS_IDLE, STATUS_PAUSED, STATUS_EXCLUDED, STATUS_ERROR]:
        row = {"status": status, "is_official_project": True}
        assert format_activity_project_cell(row) == "—"


@pytest.mark.parametrize("status", [STATUS_IDLE, STATUS_PAUSED, STATUS_ERROR])
def test_non_normal_status_does_not_display_project(temp_db, status):
    aid = _create_activity(start="09:00:00", end="09:10:00", status=status)
    csv_rows = export_service.build_statistics_csv_rows(TODAY, TODAY)
    assert csv_rows == []


def test_excluded_in_history_stats_export_but_no_project_column(temp_db):
    aid = _create_activity(start="09:00:00", end="09:10:00", status=STATUS_EXCLUDED)

    # In export
    csv_rows = export_service.build_statistics_csv_rows(TODAY, TODAY)
    assert csv_rows == []

    # In statistics total_duration
    summary = statistics_service.get_statistics_export_summary(TODAY, TODAY)
    assert summary["total_duration_seconds"] == 600
    # Not in by_project
    by_project_names = [g["display_name"] for g in summary["by_project"]]
    assert all(name == UNCATEGORIZED_PROJECT or name != "—" for name in by_project_names)


def test_classified_duration_counts_report_visible_projects(temp_db):
    """classified_duration = report-visible project duration."""
    official_pid = project_service.create_project("OfficialA")
    context_pid = project_service.create_project("ContextA")
    # Official activity (manual)
    _create_activity(
        app="Word",
        title="A1.docx",
        start="09:00:00",
        end="09:10:00",
        project_id=official_pid,
    )
    # Non-official activity (suggested)
    aid2 = _create_activity(
        app="Word",
        title="B1.docx",
        start="09:10:00",
        end="09:20:00",
    )
    _set_assignment_source(
        aid2,
        "suggested_project_name",
        project_id=_uncategorized_id(),
        suggested_name="SuggestedB",
    )
    aid3 = _create_activity(
        app="Edge",
        title="Context",
        start="09:20:00",
        end="09:30:00",
    )
    _set_assignment_source(aid3, "same_project_context", project_id=context_pid)

    summary = statistics_service.get_summary(TODAY, TODAY)
    # classified = 1200 (official + report-visible context)
    assert summary["classified_duration"] == 1200
    # uncategorized = 600 (the suggested activity)
    assert summary["uncategorized_duration"] == 600
    # total = 1800
    assert summary["total_duration"] == 1800


def test_details_project_fields_do_not_show_candidate(temp_db):
    """get_session_activity_details must not show candidate project as formal."""
    aid = _create_activity(start="09:00:00", end="09:10:00")
    _set_assignment_source(
        aid,
        "suggested_project_name",
        project_id=_uncategorized_id(),
        suggested_name="CandidateClient",
    )

    details = timeline_service.get_session_activity_details([aid], TODAY)
    assert details[0]["project_name"] == UNCATEGORIZED_PROJECT
    assert details[0]["official_project_name"] == UNCATEGORIZED_PROJECT
    assert details[0]["is_official_project"] is False
    # candidate_project_name is retained as metadata
    assert details[0].get("candidate_project_name") == "CandidateClient"


def test_details_project_fields_show_report_visible_context(temp_db):
    pid = project_service.create_project("DerivedInternal")
    aid = _create_activity(start="09:00:00", end="09:10:00")
    _set_assignment_source(aid, "anchor_context", project_id=pid)

    details = timeline_service.get_session_activity_details([aid], TODAY)
    assert details[0]["project_name"] == "DerivedInternal"
    assert details[0]["official_project_name"] == UNCATEGORIZED_PROJECT
    assert details[0]["is_official_project"] is False
    assert details[0]["is_report_project"] is True
    assert details[0]["report_attribution_kind"] == "report_context_short_gap"


def test_details_view_model_row_carries_report_attribution_flags(temp_db):
    pid = project_service.create_project("DerivedInternal")
    aid = _create_activity(start="09:00:00", end="09:10:00")
    _set_assignment_source(aid, "anchor_context", project_id=pid)

    model = view_model_service.get_session_details_view_model([aid], TODAY)
    detail = model["activities"][0]

    assert detail["project_id"] == pid
    assert detail["project_name"] == "DerivedInternal"
    assert detail["project_description"] == ""
    assert detail["display_project"]["name"] == "DerivedInternal"
    assert detail["is_report_project"] is True
    assert detail["is_report_classified"] is True
    assert detail["is_report_uncategorized"] is False
    assert detail["is_official_project"] is False
    assert detail["is_classified"] is True
    assert detail["is_uncategorized"] is False
    assert detail["report_attribution_kind"] == "report_context_short_gap"


def test_non_official_candidate_with_prior_confirmed_stays_uncategorized(temp_db):
    """A non-official candidate never inherits a formal display project."""
    confirmed = ProjectLabel(
        name="ConfirmedProject",
        id=10,
        source="manual",
    )
    candidate = ProjectLabel(
        name="SuggestedNew",
        id=None,
        source="suggested_project_name",
        is_suggested_project=True,
    )
    prior_state = ProjectOwnershipState(
        display_project=confirmed,
        candidate_project=confirmed,
        last_confirmed_project=confirmed,
    )
    state = begin_ownership_for_new_resource(prior_state, candidate, "2026-07-01 09:00:00")
    assert state.display_project.is_uncategorized is True
    # Candidate is retained
    assert state.candidate_project.name == "SuggestedNew"
    # No pending transition (only official candidates create transitions)
    assert state.project_transition.pending is False
    # last_confirmed is unchanged
    assert state.last_confirmed_project.name == "ConfirmedProject"


def test_official_candidate_applies_immediately_without_transition(temp_db):
    """An official candidate becomes the formal display project immediately."""
    confirmed = ProjectLabel(name="ProjectA", id=10, source="manual")
    candidate = ProjectLabel(name="ProjectB", id=20, source="folder_rule")
    prior_state = ProjectOwnershipState(
        display_project=confirmed,
        candidate_project=confirmed,
        last_confirmed_project=confirmed,
    )
    state = begin_ownership_for_new_resource(prior_state, candidate, "2026-07-01 09:00:00")
    assert state.project_transition.pending is False
    assert state.project_transition.threshold_seconds == 0
    assert state.display_project.name == "ProjectB"
    assert state.candidate_project.name == "ProjectB"

    state = advance_ownership(state, "2026-07-01 09:00:35")
    assert state.display_project.name == "ProjectB"
    assert state.last_confirmed_project.name == "ProjectB"
    assert state.project_transition.pending is False


def test_is_project_anchor_requires_official(temp_db):
    """_is_project_anchor returns True only for official project anchors."""
    from worktrace.services.timeline_service import _is_project_anchor

    pid = project_service.create_project("AnchorProject")
    aid = _create_activity(
        app="Word",
        title="Contract.docx - Word",
        start="09:00:00",
        end="09:10:00",
        file_path_hint="D:\\Anchor\\Contract.docx",
        project_id=pid,
    )

    rows = timeline_service.get_report_activity_rows(TODAY, TODAY)
    assert len(rows) == 1
    # manual assignment via project_id → official
    assert _is_project_anchor(rows[0]) is True
    assert rows[0]["is_official_project"] is True


def test_is_project_anchor_false_for_midnight_anchor(temp_db):
    """midnight_anchor is NOT a project anchor for UI/report merge."""
    from worktrace.services.timeline_service import _is_project_anchor

    pid = project_service.create_project("MidnightProject")
    aid = _create_activity(
        app="Word",
        title="Report.docx - Word",
        start="09:00:00",
        end="09:10:00",
        file_path_hint="D:\\Case\\Report.docx",
    )
    _set_assignment_source(aid, "midnight_anchor", project_id=pid)

    rows = timeline_service.get_report_activity_rows(TODAY, TODAY)
    assert len(rows) == 1
    assert _is_project_anchor(rows[0]) is False
