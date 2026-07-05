"""Product-semantics tests for the live activity lifecycle / display model /
frontend live runtime.

Drives the public surface (``CollectorStateMachine`` + ``WebViewBridge``) and
asserts on product-level invariants, not private function names. Covers the
eight mandatory scenarios from the maintenance governance spec:

1. Fresh start ``<30s`` — same live identity across views; base 0; no DB row.
2. Initial short then formal — initial short drops; no pending residue.
3. Formal anchor then shorts — merge without boundary; no merge across one.
4. Pause — finalize first; Current and Recent both express paused status.
5. Idle / excluded / error — hard boundary; status-only row; KPI excluded.
6. Stop / restart / recovery — no inheritance of stale project / pending.
7. ``virtual_pending`` -> ``persisted_open`` — stable identity; no rollback.
8. Frontend ticker — DOM-only; base 0; incompatibility requests refresh.
"""

from __future__ import annotations

import json
import os
import sys
from datetime import datetime

import pytest

pytestmark = [
    pytest.mark.contract,
    pytest.mark.integration,
    pytest.mark.db,
    pytest.mark.live_display,
]

from openpyxl import load_workbook

from worktrace.collector.state_machine import CollectorStateMachine
from worktrace.constants import TIME_FORMAT
from worktrace.platforms.base import ActiveWindow
from worktrace.services import (
    activity_service,
    export_service,
    recovery_service,
    session_boundary_service,
    settings_service,
    statistics_service,
    timeline_service,
)
from worktrace.webview_ui.bridge import WebViewBridge

_HERE = os.path.dirname(os.path.abspath(__file__))
_WEBVIEW_HERE = os.path.join(_HERE, "webview")
if _WEBVIEW_HERE not in sys.path:
    sys.path.insert(0, _WEBVIEW_HERE)

from static_helpers import read_js, func_body  # type: ignore

_REPORT_DATE = "2026-06-18"


def _normal(title: str) -> ActiveWindow:
    return ActiveWindow(title, f"{title.lower()}.exe", title)


def _rows(date: str = _REPORT_DATE):
    return activity_service.get_activities_by_date(date)


def _snapshot():
    raw = settings_service.get_setting("current_activity_snapshot", "") or ""
    if not raw:
        return {}
    try:
        value = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return value if isinstance(value, dict) else {}


def _patch_today(monkeypatch, date: str = _REPORT_DATE):
    """Pin ``timeline_service.get_default_report_date`` to a fixed date so
    bridge ViewModels treat ``date`` as today."""
    monkeypatch.setattr(
        timeline_service, "get_default_report_date", lambda: date
    )


def _live_identity(state_payload: dict) -> dict:
    """Extract the unified live-identity fields shared across views."""
    clock = state_payload.get("live_clock") or {}
    current = state_payload.get("current_activity") or {}
    return {
        "display_span_id": str(clock.get("display_span_id") or current.get("display_span_id") or ""),
        "stable_live_key_hash": str(clock.get("stable_live_key_hash") or current.get("stable_live_key_hash") or ""),
        "live_state": str(clock.get("live_state") or current.get("live_state") or ""),
        "sample_id": str(state_payload.get("sample_id") or ""),
    }


def _assert_single_live_identity(*payloads) -> None:
    """All payloads MUST share the same live identity (display_span_id +
    stable_live_key_hash). Used to lock the "same backend live runtime"
    contract across Current / Recent / Timeline / Details / RefreshState."""
    ids = [_live_identity(p) for p in payloads]
    spans = {i["display_span_id"] for i in ids}
    hashes = {i["stable_live_key_hash"] for i in ids}
    assert len(spans) <= 1, (
        "display_span_id must be identical across views; got: " + repr(spans)
    )
    assert len(hashes) <= 1, (
        "stable_live_key_hash must be identical across views; got: " + repr(hashes)
    )


@pytest.mark.parametrize("elapsed", [10, 20, 29])
def test_fresh_start_under_30s_shares_live_identity_across_views(
    temp_db, monkeypatch, elapsed
):
    """A running ``<30s`` normal activity MUST surface the same live
    identity (display_span_id / stable_live_key_hash) across Current,
    Recent, Timeline, Details and the refresh-state payload. There is
    exactly one backend live runtime; pages must not invent their own."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to(
        "recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:{elapsed:02d}"
    )

    assert _rows() == [], "no DB row before 30s threshold"

    overview = bridge.get_overview()
    recent = bridge.get_recent_activities()
    refresh_state = bridge.get_refresh_state(_REPORT_DATE)
    timeline = bridge.get_timeline(_REPORT_DATE)

    _assert_single_live_identity(overview, recent, refresh_state, timeline)

    current = overview["current_activity"]
    assert int(current["elapsed_seconds"]) == elapsed
    assert int(current["display_base_seconds"]) == 0
    assert str(current.get("duration_semantic") or "") == "current_live"

    recent_rows = overview.get("activities") or []
    live_recent = [r for r in recent_rows if r.get("live_state") == "virtual_pending"]
    assert live_recent, "Recent must include the virtual_pending display-only row"
    assert int(live_recent[0]["duration_seconds"]) == elapsed
    assert int(live_recent[0]["display_base_seconds"]) == 0
    assert live_recent[0].get("editable") is False
    assert live_recent[0].get("exportable") is False


def test_fresh_start_under_30s_not_editable_not_exportable(temp_db, monkeypatch, tmp_path):
    """A ``<30s`` virtual_pending activity MUST NOT be editable or
    exportable. It is live display only; it has not entered history."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:20")

    overview = WebViewBridge().get_overview()
    live_recent = [r for r in overview["activities"] if r.get("live_state") == "virtual_pending"]
    assert live_recent
    assert live_recent[0].get("editable") is False
    assert live_recent[0].get("exportable") is False

    xlsx_path = export_service.export_excel(
        _REPORT_DATE, _REPORT_DATE, str(tmp_path / "out.xlsx")
    )
    assert load_workbook(xlsx_path)["Activity Logs"].max_row == 1, (
        "virtual_pending activity must NOT appear in export"
    )
    assert statistics_service.get_summary(_REPORT_DATE, _REPORT_DATE)["total_duration"] == 0


def test_fresh_start_under_30s_drop_on_end_without_anchor_no_kpi_residue(
    temp_db, monkeypatch
):
    """When a ``<30s`` activity ends with no legal anchor, it MUST drop.
    Overview KPI MUST NOT retain the dropped seconds."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:29")
    machine.transition_to("stopped", at_time=f"{_REPORT_DATE} 09:00:29")

    assert _rows() == []
    assert settings_service.get_setting("pending_short_seconds") == "0"
    overview = bridge.get_overview()
    assert int(overview["today_total_seconds"]) == 0


def test_initial_short_then_formal_activity_drops_initial_short(
    temp_db, monkeypatch
):
    """B runs 20s then switches to A; A exceeds 30s. B MUST drop. A's
    history duration MUST NOT include B's 20s. ``pending_short_seconds``
    MUST NOT residue, feed KPI, or feed display base."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    machine.transition_to("recording", _normal("B"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:20")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:01:20")
    machine.transition_to("stopped", at_time=f"{_REPORT_DATE} 09:01:30")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["window_title"] == "A"
    assert int(rows[0]["duration_seconds"]) == 70
    assert settings_service.get_setting("pending_short_seconds") == "0"
    overview = bridge.get_overview()
    assert int(overview["today_total_seconds"]) == 70


def test_formal_anchor_then_short_activities_merge_when_no_boundary(temp_db):
    """A is a legal closed/open normal anchor. B ``<30s`` and C ``<30s``
    end without a hard boundary between them and A. B/C seconds MUST
    merge into A."""
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:01:00")
    machine.transition_to("recording", _normal("B"), at_time=f"{_REPORT_DATE} 09:05:00")
    machine.transition_to("recording", _normal("C"), at_time=f"{_REPORT_DATE} 09:05:20")
    machine.transition_to("stopped", at_time=f"{_REPORT_DATE} 09:05:29")

    rows = _rows()
    assert [r["window_title"] for r in rows] == ["A"]
    assert int(rows[0]["duration_seconds"]) == 329
    assert settings_service.get_setting("pending_short_seconds") == "0"


@pytest.mark.parametrize(
    "boundary_state,boundary_label",
    [
        ("paused", "paused"),
        ("idle", "idle"),
        ("excluded", "excluded"),
        ("error", "error"),
    ],
)
def test_short_activity_does_not_merge_across_hard_boundary(
    temp_db, boundary_state, boundary_label
):
    """A ``<30s`` short activity ending AFTER a hard boundary (pause /
    idle / excluded / error) MUST NOT merge into a pre-boundary anchor."""
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:01:00")
    machine.transition_to(boundary_state, at_time=f"{_REPORT_DATE} 09:02:00")

    machine.transition_to("recording", _normal("B"), at_time=f"{_REPORT_DATE} 09:03:00")
    machine.transition_to("recording", _normal("C"), at_time=f"{_REPORT_DATE} 09:03:20")

    rows = _rows()
    a_rows = [r for r in rows if r["window_title"] == "A"]
    assert len(a_rows) == 1
    assert int(a_rows[0]["duration_seconds"]) == 120
    b_rows = [r for r in rows if r["window_title"] == "B"]
    assert b_rows == []
    assert int(settings_service.get_setting("pending_short_seconds") or 0) == 0
    boundaries = session_boundary_service.list_boundaries(
        f"{_REPORT_DATE} 09:02:00", f"{_REPORT_DATE} 09:02:00"
    )
    assert boundaries and boundaries[-1]["reason"] == boundary_label


def test_pause_boundary_current_and_recent_both_express_paused_status(
    temp_db, monkeypatch
):
    """When the user pauses, Current AND Recent MUST both express the
    paused status. The spec explicitly forbids the case where Current has
    a paused state but Recent has no status expression at all.

    The status-only row MUST have ``contributes_to_totals=False`` so KPI
    does not increase, and MUST NOT be editable / exportable."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:30")
    machine.pause(at_time=f"{_REPORT_DATE} 09:01:00")

    overview = bridge.get_overview()
    refresh_state = bridge.get_refresh_state(_REPORT_DATE)

    current = overview["current_activity"]
    assert current.get("is_paused") is True or str(current.get("status") or "") == "paused", (
        "Current must express paused status when the user pauses"
    )

    recent_rows = overview.get("activities") or []
    status_only = [r for r in recent_rows if r.get("row_kind") == "status_only"]
    assert status_only, (
        "Recent must materialize a status-only row when paused; "
        "Current and Recent status expression must be consistent"
    )
    assert status_only[0].get("status") == "paused"
    assert status_only[0].get("contributes_to_totals") is False
    assert status_only[0].get("editable") is False
    assert status_only[0].get("exportable") is False

    assert int(overview["today_total_seconds"]) == 60


def test_pause_finalizes_normal_before_entering_paused(temp_db):
    """Entering pause MUST first finalize the current normal activity.
    The normal activity MUST be closed (have an end_time) before the
    paused status row appears."""
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:01:00")
    machine.pause(at_time=f"{_REPORT_DATE} 09:02:00")

    rows = _rows()
    a_rows = [r for r in rows if r["window_title"] == "A"]
    assert len(a_rows) == 1
    assert a_rows[0]["end_time"] is not None, (
        "normal activity must be finalized (closed) before paused status"
    )
    assert int(a_rows[0]["duration_seconds"]) == 120


@pytest.mark.parametrize("state", ["idle", "excluded", "error"])
def test_idle_excluded_error_boundary_materializes_status_only_row_excluded_from_kpi(
    temp_db, monkeypatch, state
):
    """For idle / excluded / error boundaries, a status-only display row
    MAY be materialized, but it MUST NOT contribute to KPI / totals, MUST
    NOT be editable, MUST NOT be exportable, and MUST NOT masquerade as
    a normal activity."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:30")
    machine.transition_to(state, at_time=f"{_REPORT_DATE} 09:01:00")

    overview = bridge.get_overview()
    recent_rows = overview.get("activities") or []
    status_only = [r for r in recent_rows if r.get("row_kind") == "status_only"]

    for row in status_only:
        assert row.get("contributes_to_totals") is False
        assert row.get("editable") is False
        assert row.get("exportable") is False
        assert row.get("row_kind") == "status_only"
        assert row.get("status") == state

    assert int(overview["today_total_seconds"]) == 60


def test_idle_30s_does_not_merge_into_normal_short(temp_db):
    """Idle running 30s MAY form a system status record, but MUST NOT
    merge into a normal short activity. The hard boundary isolates them."""
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:01:00")
    machine.transition_to("idle", at_time=f"{_REPORT_DATE} 09:05:00")
    machine.transition_to("recording", _normal("B"), at_time=f"{_REPORT_DATE} 09:05:29")

    rows = _rows()
    a_rows = [r for r in rows if r["window_title"] == "A"]
    assert len(a_rows) == 1
    assert int(a_rows[0]["duration_seconds"]) == 300
    assert settings_service.get_setting("pending_short_seconds") == "0"


def test_stop_short_with_anchor_merges_into_anchor(temp_db):
    """On stop, a ``<30s`` normal activity MUST merge into a legal anchor
    if one exists. No pending residue."""
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:01:00")
    machine.transition_to("recording", _normal("B"), at_time=f"{_REPORT_DATE} 09:02:00")
    machine.transition_to("stopped", at_time=f"{_REPORT_DATE} 09:02:20")
    rows = _rows()
    assert [r["window_title"] for r in rows] == ["A"]
    assert int(rows[0]["duration_seconds"]) == 140
    assert settings_service.get_setting("pending_short_seconds") == "0"


def test_stop_short_without_anchor_drops(temp_db):
    """On stop, a ``<30s`` normal activity with NO legal anchor MUST drop.
    No pending residue, no DB row."""
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("Solo"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("stopped", at_time=f"{_REPORT_DATE} 09:00:29")
    assert _rows() == []
    assert settings_service.get_setting("pending_short_seconds") == "0"


def test_restart_does_not_inherit_previous_session_project_or_pending(
    temp_db, monkeypatch
):
    """After stop + restart, the first activity MUST NOT inherit the
    previous session's project ownership, pending short, or display base.

    Stale snapshot / pending state MUST be cleared by recovery before the
    new live display sample; the fresh activity's Recent / Current MUST
    NOT show old-project time pollution."""
    _patch_today(monkeypatch)

    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("OldProject"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("OldProject"), at_time=f"{_REPORT_DATE} 09:01:00")
    machine.transition_to("stopped", at_time=f"{_REPORT_DATE} 09:02:00")
    assert len(_rows()) == 1

    settings_service.set_setting("pending_short_seconds", "999")
    settings_service.set_setting(
        "current_activity_snapshot",
        json.dumps({"status": "normal", "inferred_project_name": "OldProject"}),
    )

    recovery_service.recover_unclosed_records()
    assert settings_service.get_setting("current_activity_snapshot") == ""
    assert settings_service.get_setting("pending_short_seconds") == "0"

    machine2 = CollectorStateMachine()
    bridge = WebViewBridge()
    machine2.transition_to("recording", _normal("NewProject"), at_time=f"{_REPORT_DATE} 10:00:00")
    machine2.transition_to("recording", _normal("NewProject"), at_time=f"{_REPORT_DATE} 10:00:15")

    overview = bridge.get_overview()
    current = overview["current_activity"]
    recent_live = [r for r in overview["activities"] if r.get("live_state") == "virtual_pending"]

    current_identity = str(
        current.get("resource_name")
        or current.get("app_name")
        or current.get("display")
        or ""
    )
    assert "NewProject" in current_identity, (
        "Current resource must be NewProject after restart, not stale OldProject; "
        "got: " + repr(current_identity)
    )
    assert "OldProject" not in current_identity, (
        "Current resource MUST NOT inherit OldProject from previous session"
    )
    assert recent_live, "Recent must include the fresh virtual_pending row"
    assert int(recent_live[0]["display_base_seconds"]) == 0, (
        "fresh activity display base MUST be 0; must not inherit stale pending"
    )
    assert int(recent_live[0]["duration_seconds"]) == 15

    assert int(settings_service.get_setting("pending_short_seconds") or 0) == 0


def test_startup_recovery_does_not_merge_new_session_into_old_project(temp_db):
    """Startup recovery MUST NOT merge a new session's activity into an
    old project unless the recovery conditions (continuous, same
    resource, no boundary) are explicitly met. A restart boundary blocks
    the merge."""
    from worktrace.services import session_boundary_service

    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time=f"{_REPORT_DATE} 09:01:00")
    machine.transition_to("stopped", at_time=f"{_REPORT_DATE} 09:02:00")

    session_boundary_service.record_boundary(f"{_REPORT_DATE} 09:03:00", "restart")

    machine.transition_to("recording", _normal("B"), at_time=f"{_REPORT_DATE} 09:04:00")
    machine.transition_to("recording", _normal("C"), at_time=f"{_REPORT_DATE} 09:04:20")

    rows = _rows()
    a_rows = [r for r in rows if r["window_title"] == "A"]
    assert len(a_rows) == 1
    assert int(a_rows[0]["duration_seconds"]) == 120
    b_rows = [r for r in rows if r["window_title"] == "B"]
    assert b_rows == []


def test_virtual_to_persisted_handoff_preserves_stable_live_identity(
    temp_db, monkeypatch
):
    """At the 30s handoff, the stable live identity (stable_live_key_hash /
    display_span_id) MUST remain continuous. Current duration MUST advance
    from 29 to 30 without rollback, jump, or duplicate materialization."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:29")

    state_29 = bridge.get_refresh_state(_REPORT_DATE)
    overview_29 = bridge.get_overview()
    assert _rows() == []

    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:30")

    state_30 = bridge.get_refresh_state(_REPORT_DATE)
    overview_30 = bridge.get_overview()
    rows_30 = _rows()

    assert len(rows_30) == 1

    id_29 = _live_identity(state_29)
    id_30 = _live_identity(state_30)
    assert id_29["stable_live_key_hash"] == id_30["stable_live_key_hash"], (
        "stable_live_key_hash MUST survive virtual_pending -> persisted_open handoff"
    )
    assert id_29["display_span_id"] == id_30["display_span_id"], (
        "display_span_id MUST survive the handoff for frontend continuity"
    )

    assert int(overview_29["current_activity"]["elapsed_seconds"]) == 29
    assert int(overview_30["current_activity"]["elapsed_seconds"]) == 30
    assert int(overview_30["current_activity"]["display_base_seconds"]) == 0

    live_recent = [
        r for r in overview_30.get("activities") or []
        if r.get("live_state") in ("virtual_pending", "persisted_open")
    ]
    assert len(live_recent) == 1, (
        "Recent must NOT double-materialize a live row alongside the DB row"
    )


def test_virtual_to_persisted_handoff_refresh_revision_changes_on_structure(
    temp_db, monkeypatch
):
    """The 29s -> 30s handoff is a structural change (persisted state
    flips), so ``refresh_revision`` MUST change. This is a structural
    signal, not a natural-elapsed signal."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:29")
    rev_29 = bridge.get_refresh_state(_REPORT_DATE)["refresh_revision"]

    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:30")
    rev_30 = bridge.get_refresh_state(_REPORT_DATE)["refresh_revision"]
    assert rev_29 != rev_30, "refresh_revision MUST change on persisted-state handoff"


def test_persisted_open_natural_elapsed_growth_does_not_change_refresh_revision(
    temp_db, monkeypatch
):
    """After the 30s handoff (persisted_open), natural elapsed growth
    (30s -> 31s -> 45s) MUST NOT change ``refresh_revision``. Heavy
    refresh must NOT be triggered by elapsed advance within the same
    activity. Only structural changes flip the revision."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    fixed_start = (datetime.strptime(f"{_REPORT_DATE} 09:00:00", TIME_FORMAT)).strftime(TIME_FORMAT)
    machine.transition_to("recording", _normal("Doc"), at_time=fixed_start)
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:30")

    rev_at_30 = bridge.get_refresh_state(_REPORT_DATE)["refresh_revision"]

    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:45")
    rev_at_45 = bridge.get_refresh_state(_REPORT_DATE)["refresh_revision"]

    assert rev_at_30 == rev_at_45, (
        "refresh_revision MUST NOT change on natural elapsed growth within "
        "the same persisted_open activity"
    )


def test_frontend_apply_local_ticker_does_not_call_bridge_or_db():
    """``applyLocalTicker`` MUST be DOM-only: it must NOT call the bridge
    or write the DB. It only updates DOM text from the accepted runtime
    + DOM ``data-display-base-seconds``."""
    source = read_js("core.js")
    body = func_body(source, "applyLocalTicker")
    assert body, "core.js must define applyLocalTicker"
    assert "callBridge" not in body, (
        "applyLocalTicker must not call the bridge (DOM-only projection)"
    )
    assert "App.callBridge" not in body


def test_frontend_ticker_does_not_read_structural_caches_as_live_source():
    """``applyLocalTicker`` MUST only read the accepted ``App.liveRuntime``
    and DOM ``data-display-base-seconds``. It MUST NOT read
    ``lastOverviewSnapshot`` / ``lastRecentData`` / ``lastTimelineData`` /
    ``lastSessionDetailsViewModel`` as a live-seconds source."""
    source = read_js("core.js")
    body = func_body(source, "applyLocalTicker")
    forbidden = (
        "lastOverviewSnapshot",
        "lastRecentData",
        "lastTimelineData",
        "lastSessionDetailsViewModel",
    )
    for token in forbidden:
        assert token not in body, (
            "applyLocalTicker must not read " + token + " as a live-seconds source"
        )


def test_frontend_page_payload_incompatibility_requests_refresh():
    """When a page payload is NOT compatible with the accepted runtime,
    the frontend MUST request a refresh (set the refresh-requested flag)
    and MUST NOT mix the stale runtime with the new payload."""
    source = read_js("core.js")
    accept_body = func_body(source, "acceptPagePayloadRuntime")
    assert "isPagePayloadCompatibleWithRuntime" in accept_body
    assert "noteRejectedPagePayload" in accept_body
    reject_body = func_body(source, "noteRejectedPagePayload")
    assert "liveClockContractRefreshRequested" in reject_body, (
        "noteRejectedPagePayload must set App.liveClockContractRefreshRequested "
        "so the heartbeat knows to trigger a heavy refresh"
    )


def test_frontend_current_live_target_base_zero_aggregate_uses_static_base():
    """Current live target's ``data-display-base-seconds`` MUST be 0.
    Recent / Timeline / Details / KPI live targets MUST use their own
    static display base + the same current elapsed (single runtime).

    Verified statically: core.js renders the current-activity surface
    with ``data-duration-semantic="current-live"`` and a runtime guard
    that rejects non-zero bases for current-live targets; overview.js /
    timeline.js render aggregate rows with ``aggregate-live`` semantic
    and their own ``data-display-base-seconds``."""
    core_src = read_js("core.js")
    overview_src = read_js("overview.js")
    timeline_src = read_js("timeline.js")
    assert 'data-duration-semantic="current-live"' in core_src, (
        "core.js must tag the current-activity surface with "
        'data-duration-semantic="current-live"'
    )
    assert "current_live_target_nonzero_base" in core_src, (
        "core.js must guard current-live targets against non-zero bases"
    )
    assert (
        "aggregate-live" in overview_src or "aggregate-live" in timeline_src
    ), "Recent/Timeline aggregate rows must use aggregate-live semantic"
    assert "data-display-base-seconds" in overview_src
    assert "data-display-base-seconds" in timeline_src


def test_historical_date_no_tickable_live_clock_on_any_view(temp_db, monkeypatch):
    """On a historical (non-today) page, NO view may register a tickable
    live clock. The current open activity's live seconds MUST NOT
    pollute the historical Timeline / Details / Recent / KPI."""
    _patch_today(monkeypatch, date="2026-06-18")

    machine = CollectorStateMachine()
    bridge = WebViewBridge()
    machine.transition_to("recording", _normal("Today"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("Today"), at_time="2026-06-18 09:00:30")

    historical = "2026-06-17"
    timeline = bridge.get_timeline(historical)
    overview_historical = bridge.get_overview()

    timeline_clock = (timeline.get("live_clock") or {})
    assert timeline_clock.get("is_live") is False
    assert timeline_clock.get("is_project_duration_live") is False
    assert str(timeline_clock.get("live_state") or "") == "none"
    assert timeline_clock.get("display_span_id") == ""

    for session in timeline.get("sessions") or []:
        assert session.get("live_state") != "persisted_open"
        assert session.get("live_state") != "virtual_pending"

    stats = statistics_service.get_summary(historical, historical)
    assert stats["total_duration"] == 0


def test_single_sample_contract_views_share_same_sample_id(temp_db, monkeypatch):
    """Overview / Recent / Timeline / Details / RefreshState MUST be
    built from the same single snapshot sample. The ``sample_id`` (or
    ``stable_live_key_hash`` when present) MUST be identical across
    views fetched in the same wall-clock window."""
    _patch_today(monkeypatch)
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:00")
    machine.transition_to("recording", _normal("Doc"), at_time=f"{_REPORT_DATE} 09:00:30")

    overview = bridge.get_overview()
    recent = bridge.get_recent_activities()
    refresh_state = bridge.get_refresh_state(_REPORT_DATE)
    timeline = bridge.get_timeline(_REPORT_DATE)

    sample_ids = {
        overview.get("sample_id"),
        recent.get("sample_id"),
        refresh_state.get("sample_id"),
        timeline.get("sample_id"),
    }
    sample_ids.discard(None)
    sample_ids.discard("")
    assert len(sample_ids) <= 1, (
        "All live-bearing views must share the same sample_id (single-sample "
        "contract); got: " + repr(sample_ids)
    )


def test_pending_short_seconds_does_not_feed_display_base_or_kpi(temp_db, monkeypatch):
    """``pending_short_seconds`` MUST NOT participate in display base,
    KPI base, refresh revision, or Recent/Timeline projection. It may
    only exist as a zeroed compatibility cleanup target."""
    _patch_today(monkeypatch)
    settings_service.set_setting("pending_short_seconds", "999")
    settings_service.clear_settings_cache()

    bridge = WebViewBridge()
    overview = bridge.get_overview()
    refresh_state = bridge.get_refresh_state(_REPORT_DATE)

    assert int(overview.get("today_total_seconds") or 0) == 0
    for row in overview.get("activities") or []:
        assert int(row.get("display_base_seconds") or 0) != 999
        assert int(row.get("aggregate_display_base_seconds") or 0) != 999
    assert "pending_short_seconds" not in json.dumps(refresh_state)


def test_pending_short_seconds_static_only_zeroed_or_cleanup(temp_db):
    """``pending_short_seconds`` may only be set to 0 by the recorder /
    recovery paths. No business path may leave a non-zero value that
    survives a boundary. This is a static guard against the legacy
    carry path returning as a business state."""
    for boundary_fn in [
        lambda m: m.transition_to("stopped", at_time=f"{_REPORT_DATE} 09:00:00"),
        lambda m: m.pause(at_time=f"{_REPORT_DATE} 09:00:00"),
    ]:
        settings_service.set_setting("pending_short_seconds", "55")
        settings_service.set_setting(
            "current_activity_snapshot", '{"status":"normal"}'
        )
        machine = CollectorStateMachine()
        boundary_fn(machine)
        assert settings_service.get_setting("pending_short_seconds") == "0", (
            "pending_short_seconds MUST be zeroed by every hard boundary"
        )
