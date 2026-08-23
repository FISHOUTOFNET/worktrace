from __future__ import annotations

import os
import tempfile

import pytest
from openpyxl import load_workbook

from tests.support import activity_factory as activity_service
from worktrace.constants import STATUS_ERROR, STATUS_EXCLUDED, STATUS_IDLE, STATUS_PAUSED
from worktrace.db import get_connection, now_str
from worktrace.exports.excel_exporter import export_excel_file
from worktrace.services import project_service

pytestmark = [pytest.mark.db, pytest.mark.contract]


def _insert_closed_activity(
    app_name: str,
    process_name: str,
    window_title: str,
    start: str,
    end: str,
    duration: int,
    status: str = "normal",
    file_path_hint: str | None = None,
    project_id: int | None = None,
) -> int:
    aid = activity_service.create_activity(
        app_name,
        process_name,
        window_title,
        status=status,
        start_time=start,
        file_path_hint=file_path_hint,
        project_id=project_id,
        manual_override=project_id is not None,
    )
    activity_service.close_activity(aid, end, duration_seconds=duration)
    if project_id is not None:
        ts = now_str()
        with get_connection() as conn:
            conn.execute(
                """
                INSERT INTO activity_project_assignment(
                    activity_id, project_id, confidence, source, is_manual, created_at, updated_at
                )
                VALUES (?, ?, 100, 'manual', 1, ?, ?)
                ON CONFLICT(activity_id) DO UPDATE SET
                    project_id = excluded.project_id,
                    source = excluded.source,
                    is_manual = excluded.is_manual,
                    updated_at = excluded.updated_at
                """,
                (aid, project_id, ts, ts),
            )
    return aid


def _sessions_sheet(path: str):
    wb = load_workbook(path)
    assert "Sessions" in wb.sheetnames
    assert "Activity Logs" not in wb.sheetnames
    return wb["Sessions"]


def test_excel_sessions_sheet_omits_internal_and_resource_detail_columns(temp_db):
    _insert_closed_activity(
        "Word",
        "winword.exe",
        "合同.docx - Word",
        "2026-06-18 09:00:00",
        "2026-06-18 09:30:00",
        1800,
        file_path_hint="D:\\Docs\\合同.docx",
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test.xlsx")
        export_excel_file("2026-06-18", "2026-06-18", path)
        ws = _sessions_sheet(path)
        headers = [cell.value for cell in ws[1]]
        for removed in (
            "应用",
            "资源类型",
            "资源名称",
            "路径",
            "域名",
            "结束时间",
            "时长秒数",
            "状态",
            "修正时长",
            "是否已修正",
        ):
            assert removed not in headers
        assert headers == [
            "日期",
            "开始时间",
            "时长",
            "项目",
            "备注",
        ]


def test_excel_sessions_project_display_contract_without_internal_status(temp_db):
    pid = project_service.create_project("Client")
    _insert_closed_activity(
        "Word",
        "winword.exe",
        "合同.docx - Word",
        "2026-06-18 09:00:00",
        "2026-06-18 09:30:00",
        1800,
        project_id=pid,
        file_path_hint="D:\\Docs\\合同.docx",
    )
    for status, app, start, end in (
        (STATUS_IDLE, "空闲", "2026-06-18 09:30:00", "2026-06-18 09:40:00"),
        (STATUS_PAUSED, "已暂停", "2026-06-18 09:40:00", "2026-06-18 09:45:00"),
        (STATUS_EXCLUDED, "已排除", "2026-06-18 09:45:00", "2026-06-18 09:50:00"),
        (STATUS_ERROR, "异常", "2026-06-18 09:50:00", "2026-06-18 09:55:00"),
    ):
        _insert_closed_activity(
            app,
            status,
            f"{app}窗口",
            start,
            end,
            300,
            status=status,
            project_id=pid,
            file_path_hint="D:\\Secret\\should-not-export.txt",
        )

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test.xlsx")
        export_excel_file("2026-06-18", "2026-06-18", path)
        ws = _sessions_sheet(path)
        headers = [cell.value for cell in ws[1]]
        assert "状态" not in headers
        project_col = headers.index("项目")
        rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]

        assert len(rows) == 2
        assert {row[project_col] for row in rows} == {"Client"}
        exported_text = "\n".join(str(cell) for row in rows for cell in row)
        for raw in ("normal", "idle", "paused", "excluded", "error"):
            assert raw not in exported_text
