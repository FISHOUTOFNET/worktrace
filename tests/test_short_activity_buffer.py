import json

import pytest

pytestmark = [pytest.mark.collector_runtime, pytest.mark.integration, pytest.mark.db]

from openpyxl import load_workbook

from worktrace.collector.state_machine import CollectorStateMachine
from worktrace.constants import HISTORY_PERSIST_THRESHOLD_SECONDS
from worktrace.platforms.base import ActiveWindow
from worktrace.services import activity_service, export_service, session_boundary_service, settings_service, statistics_service
from worktrace.webview_ui.bridge import WebViewBridge


def _rows():
    return activity_service.get_activities_by_date("2026-06-18")


def _snapshot():
    return json.loads(settings_service.get_setting("current_activity_snapshot", "") or "{}")


def _normal(title: str) -> ActiveWindow:
    return ActiveWindow(title, f"{title.lower()}.exe", title)


def test_single_auto_activity_29_seconds_has_snapshot_but_no_history_stats_or_export(temp_db, tmp_path):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("Doc"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("Doc"), at_time="2026-06-18 09:00:29")

    assert _snapshot()["window_title"] == "Doc"
    assert _snapshot()["is_persisted"] is False
    assert _rows() == []
    assert statistics_service.get_summary("2026-06-18", "2026-06-18")["total_duration"] == 0

    xlsx_path = export_service.export_excel("2026-06-18", "2026-06-18", str(tmp_path / "out.xlsx"))
    assert load_workbook(xlsx_path)["Activity Logs"].max_row == 1


def test_29_and_30_second_threshold_updates_snapshot_ui_and_revision_same_tick(temp_db):
    settings_service.set_setting("collector_status", "running")
    settings_service.set_setting("user_paused", "false")
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    machine.transition_to("recording", _normal("Doc"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("Doc"), at_time="2026-06-18 09:00:29")

    snap_29 = _snapshot()
    rows_29 = _rows()
    state_29 = bridge.get_refresh_state("2026-06-18")
    overview_29 = bridge.get_overview()

    assert rows_29 == []
    assert snap_29["is_persisted"] is False
    assert snap_29["persisted_activity_id"] is None
    assert state_29["is_persisted"] is False
    assert state_29["persisted_activity_id"] == 0
    assert "暂不入历史" in overview_29["current_activity"]["display"]

    machine.transition_to("recording", _normal("Doc"), at_time="2026-06-18 09:00:30")

    snap_30 = _snapshot()
    rows_30 = _rows()
    state_30 = bridge.get_refresh_state("2026-06-18")
    overview_30 = bridge.get_overview()

    assert len(rows_30) == 1
    assert rows_30[0]["start_time"] == "2026-06-18 09:00:00"
    assert rows_30[0]["end_time"] is None
    assert snap_30["is_persisted"] is True
    assert snap_30["persisted_activity_id"] == rows_30[0]["id"]
    assert state_30["is_persisted"] is True
    assert state_30["persisted_activity_id"] == rows_30[0]["id"]
    assert "已进入历史" in overview_30["current_activity"]["display"]
    assert state_29["refresh_revision"] != state_30["refresh_revision"]


def test_history_persist_threshold_is_shared_constant():
    """The threshold constant is defined in ``worktrace.constants`` and
    the lifecycle facade (``activity_lifecycle_service``) enforces it.
    The recorder no longer imports its own copy; it delegates to the
    facade's public ``persist_open_activity_if_ready`` entry which
    checks the threshold internally."""
    from worktrace.constants import HISTORY_PERSIST_THRESHOLD_SECONDS as facade_threshold

    assert facade_threshold == HISTORY_PERSIST_THRESHOLD_SECONDS
    assert HISTORY_PERSIST_THRESHOLD_SECONDS == 30


def test_single_auto_activity_30_seconds_persists_once_with_actual_start(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("Doc"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("Doc"), at_time="2026-06-18 09:00:30")
    machine.transition_to("recording", _normal("Doc"), at_time="2026-06-18 09:00:45")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["start_time"] == "2026-06-18 09:00:00"
    assert rows[0]["end_time"] is None
    assert _snapshot()["is_persisted"] is True


def test_short_activity_merges_into_previous_formal_normal_activity(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:05:00")
    machine.transition_to("recording", _normal("C"), at_time="2026-06-18 09:05:20")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["window_title"] == "A"
    assert rows[0]["duration_seconds"] == 320


def test_same_activity_resumes_after_absorbed_short_activity_without_new_record(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:05:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:05:20")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:07:00")
    machine.transition_to("stopped", at_time="2026-06-18 09:07:00")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["window_title"] == "A"
    assert rows[0]["start_time"] == "2026-06-18 09:00:00"
    assert rows[0]["end_time"] == "2026-06-18 09:07:00"
    assert rows[0]["duration_seconds"] == 7 * 60


def test_multiple_short_activities_merge_into_previous_formal_activity(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:30")
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:05:00")
    machine.transition_to("recording", _normal("C"), at_time="2026-06-18 09:05:20")
    machine.transition_to("recording", _normal("D"), at_time="2026-06-18 09:05:29")
    machine.transition_to("stopped", at_time="2026-06-18 09:05:29")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["duration_seconds"] == 329


def test_initial_under_30_without_anchor_drops_no_pending_no_kpi_base(temp_db, monkeypatch):
    from worktrace.services import timeline_service

    monkeypatch.setattr(timeline_service, "get_default_report_date", lambda: "2026-06-18")
    machine = CollectorStateMachine()
    bridge = WebViewBridge()
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:20")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:20")
    machine.transition_to("stopped", at_time="2026-06-18 09:01:20")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["window_title"] == "A"
    assert rows[0]["duration_seconds"] == 60
    assert settings_service.get_setting("pending_short_seconds") == "0"
    assert bridge.get_overview()["today_total_seconds"] == 60


def test_initial_dropped_seconds_are_not_added_when_formal_activity_closes(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:20")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:20")
    machine.transition_to("stopped", at_time="2026-06-18 09:01:30")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["window_title"] == "A"
    assert rows[0]["duration_seconds"] == 70
    assert settings_service.get_setting("pending_short_seconds") == "0"


def test_persisted_current_activity_continues_to_90_seconds_without_duplicate_insert(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:30")
    assert len(_rows()) == 1
    machine.transition_to("idle", at_time="2026-06-18 09:01:30")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["duration_seconds"] == 90


def test_stop_short_current_activity_merges_or_pends(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:00:00")
    machine.transition_to("stopped", at_time="2026-06-18 09:00:29")
    assert _rows() == []
    assert settings_service.get_setting("pending_short_seconds") == "0"

    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:30")
    machine.transition_to("stopped", at_time="2026-06-18 09:01:30")
    assert _rows()[0]["duration_seconds"] == 30


def test_short_idle_polling_does_not_create_history(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("idle", at_time="2026-06-18 09:00:00")
    machine.transition_to("idle", at_time="2026-06-18 09:00:01")
    assert _rows() == []
    assert _snapshot()["status"] == "idle"


def test_idle_30_seconds_creates_one_idle_record(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("idle", at_time="2026-06-18 09:00:00")
    for second in range(1, 31):
        machine.transition_to("idle", at_time=f"2026-06-18 09:{second // 60:02d}:{second % 60:02d}")
        assert len(_rows()) <= 1

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["status"] == "idle"
    assert rows[0]["end_time"] is None


def test_short_idle_does_not_merge_into_previous_normal(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("idle", at_time="2026-06-18 09:05:00")
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:05:29")

    rows = _rows()
    assert len(rows) == 1
    assert rows[0]["window_title"] == "A"
    assert rows[0]["duration_seconds"] == 300
    assert settings_service.get_setting("pending_short_seconds") == "0"


def test_cross_boundary_stale_pending_not_used_as_display_base(temp_db, monkeypatch):
    """``pending_short_seconds`` is a zeroed cleanup target. Startup
    recovery clears any stale value before the next live display sample;
    the display base MUST remain 0."""
    from worktrace.services import recovery_service, timeline_service

    monkeypatch.setattr(
        timeline_service, "get_default_report_date", lambda: "2026-06-18"
    )
    settings_service.set_setting("current_activity_snapshot", '{"old": true}')
    settings_service.set_setting("pending_short_seconds", "5")

    recovery_service.recover_unclosed_records()

    assert settings_service.get_setting("current_activity_snapshot") == ""
    assert settings_service.get_setting("pending_short_seconds") == "0"

    machine = CollectorStateMachine()
    bridge = WebViewBridge()
    machine.transition_to("recording", _normal("Fresh"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("Fresh"), at_time="2026-06-18 09:00:10")

    overview = bridge.get_overview()
    current = overview["current_activity"]
    recent = overview["activities"][0]
    assert current["elapsed_seconds"] == 10
    assert recent["duration_seconds"] == 10
    assert recent["display_base_seconds"] == 0

    machine.transition_to("recording", _normal("Short"), at_time="2026-06-18 09:00:10")
    machine.transition_to("recording", _normal("Next"), at_time="2026-06-18 09:00:20")
    assert int(settings_service.get_setting("pending_short_seconds") or 0) == 0


def test_pause_resume_virtual_pending_materializes_recent_row(temp_db, monkeypatch):
    """After pause/resume, a fresh unpersisted activity must appear in the
    full Overview ViewModel so Recent is refreshed along with Current."""
    from worktrace.services import timeline_service

    monkeypatch.setattr(
        timeline_service, "get_default_report_date", lambda: "2026-06-18"
    )
    machine = CollectorStateMachine()
    bridge = WebViewBridge()

    machine.transition_to("recording", _normal("Before"), at_time="2026-06-18 09:00:00")
    machine.pause(at_time="2026-06-18 09:00:10")
    machine.transition_to("recording", _normal("After"), at_time="2026-06-18 09:01:00")
    machine.transition_to("recording", _normal("After"), at_time="2026-06-18 09:01:09")

    state = bridge.get_refresh_state("2026-06-18")
    overview = bridge.get_overview()

    assert state["current_activity"]["active"] is True
    assert state["live_clock"]["live_state"] == "virtual_pending"
    assert overview["current_activity"]["elapsed_seconds"] == 9
    assert overview["activities"], "Recent must include the display-only live row"
    assert overview["activities"][0]["live_state"] == "virtual_pending"
    assert overview["activities"][0]["duration_seconds"] == 9


# Section 四: short-activity boundary tests. Running ``absorbed_pending``
# display projection is display-only (verified in
# ``test_live_display_project_transition_contract.py``). Finished merge
# MUST NOT cross a session boundary (pause / stopped / restart / midnight).


def test_short_activity_after_restart_does_not_merge_into_pre_boundary_activity(temp_db):
    """Section 四: a ``<30s`` short activity that ends AFTER a restart
    boundary MUST NOT merge into a pre-boundary confirmed normal activity.
    The merge target lookup uses
    ``activity_service.get_latest_closed_auto_normal_activity(after_time=
    latest_boundary_time())`` so any anchor whose ``end_time`` is before
    or equal to the latest boundary is excluded. With no post-boundary
    anchor, the short seconds MUST drop and ``pending_short_seconds``
    MUST remain zero.
    """
    from worktrace.services import session_boundary_service

    # Pre-boundary anchor A (persisted, 60s, then closed by stopping at 09:02:00).
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("stopped", at_time="2026-06-18 09:02:00")

    # Simulate a restart boundary at 09:03:00 (recorded by recovery_service on app restart).
    session_boundary_service.record_boundary("2026-06-18 09:03:00", "restart")

    # After the restart, start a NEW short activity B (<30s) and finish it
    # by switching to C — this triggers the finished-short merge/drop path.
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:04:00")
    machine.transition_to("recording", _normal("C"), at_time="2026-06-18 09:04:20")

    rows = _rows()
    a_rows = [r for r in rows if r["window_title"] == "A"]
    assert len(a_rows) == 1, "anchor A must be closed and persisted"
    # A: 09:00:00 → 09:02:00 = 120s. B's 20s must NOT be added across the
    # restart boundary (the latest boundary is 09:03:00, A.end_time is 09:02:00).
    assert int(a_rows[0]["duration_seconds"]) == 120, (
        "B's <30s seconds must NOT merge into A across the restart boundary"
    )
    # B is not persisted and does not appear as a separate row.
    b_rows = [r for r in rows if r["window_title"] == "B"]
    assert b_rows == [], "B must NOT be persisted as a separate row"
    assert int(settings_service.get_setting("pending_short_seconds") or 0) == 0


def test_short_activity_after_pause_does_not_merge_into_pre_boundary_activity(temp_db):
    """Section 四: a ``<30s`` short activity that ends AFTER a pause
    boundary MUST NOT merge into a pre-boundary confirmed normal activity.
    ``state_machine.pause()`` records a ``"paused"`` boundary so the merge
    target lookup excludes any anchor at or before the pause time. B is
    finished by switching to C (not by stopping) so the merge/drop logic
    runs; ``pending_short_seconds`` MUST remain zero.
    """
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.pause(at_time="2026-06-18 09:02:00")

    # After the pause, start B (<30s) and finish by switching to C.
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:03:00")
    machine.transition_to("recording", _normal("C"), at_time="2026-06-18 09:03:20")

    rows = _rows()
    a_rows = [r for r in rows if r["window_title"] == "A"]
    assert len(a_rows) == 1
    # A: 09:00:00 → 09:02:00 = 120s. B's 20s must NOT be added across the pause boundary.
    assert int(a_rows[0]["duration_seconds"]) == 120
    b_rows = [r for r in rows if r["window_title"] == "B"]
    assert b_rows == [], "B must NOT be persisted as a separate row"
    assert int(settings_service.get_setting("pending_short_seconds") or 0) == 0


def test_short_activity_after_midnight_does_not_merge_into_pre_boundary_activity(temp_db):
    """Section 四: a ``<30s`` short activity that ends AFTER a midnight
    boundary MUST NOT merge into a pre-midnight confirmed normal activity.
    A ``"midnight"`` boundary is recorded so the merge target lookup
    excludes any anchor on the previous day. B is finished by switching
    to C so the merge/drop logic runs; ``pending_short_seconds`` MUST
    remain zero.
    """
    from worktrace.services import session_boundary_service

    # Day 1: anchor A (persisted, 300s, closed by stopping at midnight).
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 23:50:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 23:55:00")
    machine.transition_to("stopped", at_time="2026-06-19 00:00:00")

    # Record a "midnight" boundary at 2026-06-19 00:00:00.
    session_boundary_service.record_boundary("2026-06-19 00:00:00", "midnight")

    # Day 2: short activity B (<30s) after midnight, finished by switching to C.
    machine.transition_to("recording", _normal("B"), at_time="2026-06-19 09:00:00")
    machine.transition_to("recording", _normal("C"), at_time="2026-06-19 09:00:20")

    # Check A on day 1.
    rows_day1 = activity_service.get_activities_by_date("2026-06-18")
    a_rows = [r for r in rows_day1 if r["window_title"] == "A"]
    assert len(a_rows) == 1, "anchor A must be closed and persisted on day 1"
    # A: 23:50:00 → 00:00:00 = 600s. B's 20s must NOT be added across midnight.
    assert int(a_rows[0]["duration_seconds"]) == 600, (
        "B's <30s seconds must NOT merge into A across the midnight boundary"
    )
    # B is not persisted on either day.
    b_rows_day1 = [r for r in rows_day1 if r["window_title"] == "B"]
    rows_day2 = activity_service.get_activities_by_date("2026-06-19")
    b_rows_day2 = [r for r in rows_day2 if r["window_title"] == "B"]
    assert b_rows_day1 == [] and b_rows_day2 == [], "B must NOT be persisted as a separate row"
    assert int(settings_service.get_setting("pending_short_seconds") or 0) == 0


def test_short_activity_after_stopped_does_not_merge_into_pre_boundary_activity(temp_db):
    """Section 四: a ``<30s`` short activity that ends AFTER a stopped
    boundary MUST NOT merge into a pre-boundary confirmed normal activity.
    ``state_machine.transition_to("stopped")`` records a ``"stopped"``
    boundary so the merge target lookup excludes any anchor at or before
    the stop time. B is finished by switching to C (not by stopping) so
    the merge/drop logic runs; ``pending_short_seconds`` MUST remain zero.
    """
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("stopped", at_time="2026-06-18 09:02:00")

    # After the stopped boundary, start B (<30s) and finish by switching to C.
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:03:00")
    machine.transition_to("recording", _normal("C"), at_time="2026-06-18 09:03:20")

    rows = _rows()
    a_rows = [r for r in rows if r["window_title"] == "A"]
    assert len(a_rows) == 1
    # A: 09:00:00 → 09:02:00 = 120s. B's 20s must NOT be added across the stopped boundary.
    assert int(a_rows[0]["duration_seconds"]) == 120
    b_rows = [r for r in rows if r["window_title"] == "B"]
    assert b_rows == [], "B must NOT be persisted as a separate row"
    assert int(settings_service.get_setting("pending_short_seconds") or 0) == 0


def test_switch_under_30_with_anchor_merges_to_previous_anchor(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:02:00")
    machine.transition_to("recording", _normal("C"), at_time="2026-06-18 09:02:20")

    rows = _rows()
    assert [r["window_title"] for r in rows] == ["A"]
    assert rows[0]["duration_seconds"] == 140
    assert settings_service.get_setting("pending_short_seconds") == "0"


def test_switch_under_30_without_anchor_drops(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("C"), at_time="2026-06-18 09:00:20")

    assert _rows() == []
    assert settings_service.get_setting("pending_short_seconds") == "0"


def test_pause_under_30_with_anchor_merges_before_boundary(temp_db):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:02:00")
    machine.pause(at_time="2026-06-18 09:02:20")

    rows = _rows()
    assert [r["window_title"] for r in rows] == ["A"]
    assert rows[0]["duration_seconds"] == 140
    assert settings_service.get_setting("pending_short_seconds") == "0"
    boundaries = session_boundary_service.list_boundaries(
        "2026-06-18 09:02:20",
        "2026-06-18 09:02:20",
    )
    assert boundaries and boundaries[-1]["reason"] == "paused"


def test_pause_under_30_without_anchor_drops_no_pending(temp_db, monkeypatch):
    from worktrace.services import timeline_service

    monkeypatch.setattr(timeline_service, "get_default_report_date", lambda: "2026-06-18")
    machine = CollectorStateMachine()
    bridge = WebViewBridge()
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:00:00")
    machine.pause(at_time="2026-06-18 09:00:20")

    assert _rows() == []
    assert settings_service.get_setting("pending_short_seconds") == "0"
    overview = bridge.get_overview()
    assert overview["today_total_seconds"] == 0
    assert overview["activities"][0]["row_kind"] == "status_only"
    assert overview["activities"][0]["display_status"] == "已暂停"
    assert overview["activities"][0]["contributes_to_totals"] is False


def test_stop_shutdown_time_jump_under_30_share_same_finalizer(temp_db):
    scenarios = [
        ("StopA", "09:00:00", "09:01:00", "09:02:00", "09:02:20", lambda m, t: m.transition_to("stopped", at_time=t)),
        ("ShutdownA", "09:10:00", "09:11:00", "09:12:00", "09:12:20", lambda m, t: m._stop_recording_at_boundary(t, "shutdown")),
        ("JumpA", "09:20:00", "09:21:00", "09:22:00", "09:22:20", lambda m, t: m.reset_for_time_jump(t)),
    ]
    for title, start, persist_at, short_start, finish_at, finish in scenarios:
        settings_service.set_setting("pending_short_seconds", "0")
        machine = CollectorStateMachine()
        machine.transition_to("recording", _normal(title), at_time=f"2026-06-18 {start}")
        machine.transition_to("recording", _normal(title), at_time=f"2026-06-18 {persist_at}")
        machine.transition_to("recording", _normal(title + 'B'), at_time=f"2026-06-18 {short_start}")
        finish(machine, f"2026-06-18 {finish_at}")
        rows = _rows()
        anchor = [r for r in rows if r["window_title"] == title][-1]
        assert anchor["duration_seconds"] == 140
        assert settings_service.get_setting("pending_short_seconds") == "0"


@pytest.mark.parametrize("state", ["idle", "excluded", "error"])
def test_normal_to_idle_excluded_error_under_30_finalize_before_boundary(temp_db, state):
    machine = CollectorStateMachine()
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:00:00")
    machine.transition_to("recording", _normal("A"), at_time="2026-06-18 09:01:00")
    machine.transition_to("recording", _normal("B"), at_time="2026-06-18 09:02:00")
    machine.transition_to(state, at_time="2026-06-18 09:02:20")

    rows = _rows()
    a_rows = [r for r in rows if r["window_title"] == "A"]
    assert len(a_rows) == 1
    assert a_rows[0]["duration_seconds"] == 140
    assert settings_service.get_setting("pending_short_seconds") == "0"
    boundaries = session_boundary_service.list_boundaries(
        "2026-06-18 09:02:20",
        "2026-06-18 09:02:20",
    )
    assert boundaries and boundaries[-1]["reason"] == state


def test_stopped_boundary_clears_stale_pending_runtime_state(temp_db):
    settings_service.set_setting("pending_short_seconds", "12")
    settings_service.set_setting("current_activity_snapshot", '{"status":"normal"}')

    machine = CollectorStateMachine()
    machine.transition_to("stopped", at_time="2026-06-18 09:00:00")

    assert settings_service.get_setting("pending_short_seconds") == "0"
    assert settings_service.get_setting("current_activity_snapshot", "") == ""


def test_paused_boundary_clears_stale_pending_runtime_state(temp_db):
    settings_service.set_setting("pending_short_seconds", "12")
    settings_service.set_setting("current_activity_snapshot", '{"status":"normal"}')

    machine = CollectorStateMachine()
    machine.pause(at_time="2026-06-18 09:00:00")

    assert settings_service.get_setting("pending_short_seconds") == "0"
    assert _snapshot()["status"] == "paused"


def test_midnight_boundary_clears_stale_pending_runtime_state(temp_db):
    settings_service.set_setting("pending_short_seconds", "12")
    settings_service.set_setting("current_activity_snapshot", '{"status":"normal"}')

    machine = CollectorStateMachine()
    machine.split_at_midnight("2026-06-19 00:00:00")

    assert settings_service.get_setting("pending_short_seconds") == "0"
    assert settings_service.get_setting("current_activity_snapshot", "") == ""
